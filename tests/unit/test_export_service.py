"""Unit tests for ExportService — خدمة التصدير (CSV/Excel/PDF)."""
import csv
import io
import sys
import types
from datetime import date, datetime
from decimal import Decimal
from types import SimpleNamespace

from openpyxl import load_workbook

from services.export_service import ExportService


def _csv_rows(buf):
    text = buf.getvalue().decode('utf-8-sig')
    return list(csv.reader(io.StringIO(text)))


def _wb(buf):
    buf.seek(0)
    return load_workbook(buf)


def _install_pdf_stub(monkeypatch):
    captured = []

    class StubHTML:
        def __init__(self, string=None, **kwargs):
            captured.append(string)

        def write_pdf(self, *args, **kwargs):
            return b'%PDF-fake'

    monkeypatch.setitem(sys.modules, 'weasyprint',
                        types.SimpleNamespace(HTML=StubHTML))
    return captured


class TestCsvExport:
    def test_roundtrip_headers_and_rows(self):
        headers = ['الكود', 'الاسم', 'المبلغ']
        rows = [['A-1', 'فرامل', '100.500'], ['B-2', 'فلتر زيت', '42.000']]
        out = ExportService.export_to_csv(rows, headers)
        parsed = _csv_rows(out)
        assert parsed[0] == headers
        assert parsed[1] == ['A-1', 'فرامل', '100.500']
        assert parsed[2] == ['B-2', 'فلتر زيت', '42.000']

    def test_utf8_bom_present_for_excel_arabic(self):
        out = ExportService.export_to_csv([['سلام']], ['تحية'])
        raw = out.getvalue()
        assert raw[:3] == b'\xef\xbb\xbf'
        assert 'سلام'.encode('utf-8') in raw[3:]

    def test_empty_dataset_yields_headers_only(self):
        out = ExportService.export_to_csv([], ['h1', 'h2'])
        parsed = _csv_rows(out)
        assert parsed == [['h1', 'h2']]

    def test_cell_coercion_decimal_date_none_bool(self):
        row = [Decimal('12.500'), date(2026, 2, 1), None, True]
        parsed = _csv_rows(ExportService.export_to_csv([row], ['c']))
        assert parsed[1] == ['12.500', '2026-02-01', '', 'True']

    def test_commas_quotes_newlines_escaped(self):
        tricky = 'a,"b"\nsecond'
        parsed = _csv_rows(ExportService.export_to_csv([[tricky]], ['x']))
        assert parsed[1] == [tricky]

    def test_filename_params_accepted_returns_seekable_buffer(self):
        default_out = ExportService.export_to_csv([['v']], ['h'])
        named_out = ExportService.export_to_csv([['v']], ['h'], 'x.csv')
        for buf in (default_out, named_out):
            assert isinstance(buf, io.BytesIO)
            assert buf.tell() == 0


class TestExcelExport:
    HEADERS = ['Code', 'Name', 'Total']
    ROWS = [['S-1', 'Brake Pad', 1250.5], ['S-2', 'Oil Filter', 99]]

    def test_roundtrip_headers_rows_rtl(self):
        ws = _wb(ExportService.export_to_excel('تقرير المبيعات', self.HEADERS,
                                               self.ROWS)).active
        assert ws.sheet_view.rightToLeft is True
        assert ws['A1'].value == 'تقرير المبيعات'
        assert str(ws['A2'].value).startswith('Generated:')
        assert [ws.cell(row=4, column=i).value for i in (1, 2, 3)] == self.HEADERS
        assert ws['A5'].value == 'S-1' and ws['C6'].value == 99

    def test_title_truncated_to_sheet_name_limit(self):
        long_title = 'تقرير طويل جداً ' + 'x' * 40
        wb = _wb(ExportService.export_to_excel(long_title, ['H'], [['v']]))
        assert wb.active.title == long_title[:31]

    def test_ltr_mode_disables_rtl(self):
        ws = _wb(ExportService.export_to_excel(
            'Report', ['H'], [['v']], right_to_left=False)).active
        assert not ws.sheet_view.rightToLeft

    def test_numeric_cells_get_money_format_and_right_align(self):
        ws = _wb(ExportService.export_to_excel('R', ['N'], [[12.5]])).active
        cell = ws['A5']
        assert cell.number_format == '#,##0.000'
        assert cell.alignment.horizontal == 'right'

    def test_summary_section_appended_bold_with_money_format(self):
        summary = {'الإجمالي': 1500.0}
        ws = _wb(ExportService.export_to_excel(
            'R', ['H1', 'H2'], [['a', 'b']],
            summary=summary)).active
        assert ws['A7'].value == 'Summary / الإجماليات'
        label_cell = ws['A8']
        value_cell = ws['B8']
        assert label_cell.value == 'الإجمالي' and label_cell.font.b
        assert value_cell.value == 1500.0
        assert value_cell.number_format == '#,##0.000'

    def test_none_decimal_datetime_cells_coerced(self):
        stamp = datetime(2026, 3, 15, 10, 30, 0)
        ws = _wb(ExportService.export_to_excel(
            'R', ['A', 'B', 'C'],
            [[None, Decimal('1234.500'), stamp]])).active
        assert ws['A5'].value is None
        assert abs(ws['B5'].value - 1234.5) < 1e-6
        assert ws['C5'].value == stamp

    def test_empty_rows_headers_only_and_column_widths_set(self):
        ws = _wb(ExportService.export_to_excel('R', ['Code', 'N'],
                                               [])).active
        assert ws.max_row == 4
        assert ws['A4'].value == 'Code' and ws['B4'].value == 'N'
        assert ws.column_dimensions['A'].width == 14

    def test_wide_content_capped_at_40_chars(self):
        long_val = 'w' * 80
        ws = _wb(ExportService.export_to_excel(
            'R', ['Code'], [[long_val]])).active
        assert ws.column_dimensions['A'].width == 42


class TestPdfExport:
    def test_html_contains_title_headers_and_rows(self, monkeypatch):
        captured = _install_pdf_stub(monkeypatch)
        out = ExportService.export_to_pdf(
            'تقييم المخزون', ['المنتج', 'الكمية'],
            [['زيت', '10'], ['فلتر', '3.5']])
        html = captured[0]
        assert 'تقييم المخزون' in html
        assert '<th>المنتج</th>' in html and '<th>الكمية</th>' in html
        assert '<td style="text-align:right;padding:6px 8px;border:1px solid #ddd">زيت</td>' in html
        assert '{title}' not in html and '{rows_html}' not in html
        assert out.getvalue() == b'%PDF-fake'

    def test_summary_block_rendered(self, monkeypatch):
        captured = _install_pdf_stub(monkeypatch)
        ExportService.export_to_pdf('R', ['H'], [['a']],
                                    summary={'الإجمالي': '999'})
        assert 'Summary / الإجماليات' in captured[0]
        assert '<td style="font-weight:bold;padding:4px 8px">الإجمالي</td>' in captured[0]

    def test_no_summary_no_summary_heading(self, monkeypatch):
        captured = _install_pdf_stub(monkeypatch)
        ExportService.export_to_pdf('R', ['H'], [['a']])
        assert 'Summary / الإجماليات' not in captured[0]

    def test_ltr_mode_drops_rtl_attributes(self, monkeypatch):
        captured = _install_pdf_stub(monkeypatch)
        ExportService.export_to_pdf('R', ['H'], [['a']],
                                    right_to_left=False)
        html = captured[0]
        assert 'dir="rtl"' not in html
        assert 'direction:rtl;' not in html
        assert 'text-align:left' in html

    def test_real_weasyprint_render_produces_pdf_bytes(self):
        out = ExportService.export_to_pdf('Inventory', ['Item'], [['Pad', '2']])
        data = out.getvalue()
        assert data.startswith(b'%PDF')
        assert len(data) > 200


class TestDictsToRows:
    def test_selects_keys_in_given_order(self):
        items = [{'b': 2, 'a': 1, 'ignored': 3}, {'a': 4, 'b': 5}]
        rows = ExportService.dicts_to_rows(items, ['a', 'b'])
        assert rows == [[1, 2], [4, 5]]

    def test_missing_keys_default_to_empty_string(self):
        rows = ExportService.dicts_to_rows([{'a': 1}, {}], ['a', 'b'])
        assert rows == [[1, ''], ['', '']]
        assert ExportService.dicts_to_rows([], ['a']) == []


class TestLegacyPurchasesCsv:
    @staticmethod
    def _purchase(**overrides):
        base = dict(
            id=7,
            package=SimpleNamespace(name_ar='باقة أساسية'),
            customer_name='Ali Hassan',
            customer_email='ali@test.ae',
            amount_paid=Decimal('99.500'),
            payment_method='card',
            payment_status='paid',
            created_at=datetime(2026, 3, 1, 12, 30),
        )
        base.update(overrides)
        return SimpleNamespace(**base)

    def test_full_purchase_row(self):
        parsed = _csv_rows(
            ExportService.export_purchases_to_csv([self._purchase()]))
        assert parsed[0][1] == 'الباقة'
        assert parsed[1] == [
            '7', 'باقة أساسية', 'Ali Hassan', 'ali@test.ae',
            '$99.500', 'card', 'paid', '2026-03-01 12:30',
        ]

    def test_null_package_and_created_at_fall_back(self):
        parsed = _csv_rows(ExportService.export_purchases_to_csv(
            [self._purchase(package=None, created_at=None)]))
        assert parsed[1][1] == 'N/A'
        assert parsed[1][7] == 'N/A'


class TestLegacyDonationsCsv:
    def test_anonymous_donor_defaults(self):
        donation = SimpleNamespace(
            id=3, donor_name=None, donor_email=None,
            amount_usd=25, payment_method='paypal',
            status='completed', created_at=None,
        )
        parsed = _csv_rows(
            ExportService.export_donations_to_csv([donation]))
        assert parsed[1] == [
            '3', 'مجهول', 'N/A', '$25', 'paypal', 'completed', 'N/A',
        ]

    def test_named_donor_row(self):
        donation = SimpleNamespace(
            id=4, donor_name='Sara', donor_email='sara@test.ae',
            amount_usd=Decimal('10.250'), payment_method='stripe',
            status='pending', created_at=datetime(2026, 1, 5, 8, 0),
        )
        parsed = _csv_rows(
            ExportService.export_donations_to_csv([donation]))
        assert parsed[1][1:4] == ['Sara', 'sara@test.ae', '$10.250']


class TestLegacyCardsCsv:
    @staticmethod
    def _card(**overrides):
        base = dict(
            id=11,
            customer_name='Omar',
            customer_email='omar@test.ae',
            get_card_display=lambda: '**** 4242',
            card_type='visa',
            amount=55,
            status='active',
            created_at=datetime(2026, 4, 10, 9, 15),
        )
        base.update(overrides)
        return SimpleNamespace(**base)

    def test_card_row_with_display_method(self):
        parsed = _csv_rows(ExportService.export_cards_to_csv([self._card()]))
        assert parsed[1] == [
            '11', 'Omar', 'omar@test.ae', '**** 4242', 'visa',
            '$55', 'active', '2026-04-10 09:15',
        ]

    def test_missing_display_method_uses_na(self):
        card = self._card(get_card_display=None)
        del card.get_card_display
        parsed = _csv_rows(ExportService.export_cards_to_csv([card]))
        assert parsed[1][3] == 'N/A'

    def test_zero_or_none_amount_and_unknown_type(self):
        zero = self._card(amount=0, card_type=None)
        none_amt = self._card(amount=None)
        parsed = _csv_rows(
            ExportService.export_cards_to_csv([zero, none_amt]))
        assert parsed[1][4] == 'Unknown' and parsed[1][5] == 'N/A'
        assert parsed[2][5] == 'N/A'


class TestLegacyHtmlReport:
    def test_renders_title_stats_and_table(self):
        html = ExportService.generate_pdf_report('تقرير المدفوعات', {
            'stats': {'إجمالي المبيعات': 12000},
            'table_data': [['S-1', '100']],
            'table_headers': ['رقم البيع', 'المبلغ'],
        })
        assert html.count('تقرير المدفوعات') >= 2
        assert '{title}' not in html
        assert '<div class="stat-number">12000</div>' in html
        assert '<div class="stat-label">إجمالي المبيعات</div>' in html
        assert '<th>رقم البيع</th>' in html
        assert '<td>S-1</td>' in html and '<td>100</td>' in html

    def test_empty_data_omits_stats_grid_and_table(self):
        html = ExportService.generate_pdf_report('فارغ', {})
        assert '<div class="stat-box"' not in html
        assert '<table>' not in html

    def test_table_skipped_when_headers_missing_even_with_data(self):
        html = ExportService.generate_pdf_report(
            'X', {'table_data': [['a']]})
        assert '<table>' not in html
