"""
Export Service — خدمة التصدير
Export reports to Excel (.xlsx), PDF, and CSV.
Reuses openpyxl + weasyprint already in requirements.txt.
"""
from datetime import datetime, timezone
from io import BytesIO, StringIO
import csv
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """خدمة تصدير التقارير"""

    # ── CSV ──────────────────────────────────────────────────────────────

    @staticmethod
    def export_to_csv(data, headers, filename='export.csv'):
        """Export rows to a CSV file with UTF-8 BOM for Excel compat."""
        str_output = StringIO()
        writer = csv.writer(str_output)
        writer.writerow(headers)
        for row in data:
            writer.writerow(row)

        output = BytesIO()
        output.write(b'\xef\xbb\xbf')
        output.write(str_output.getvalue().encode('utf-8'))
        output.seek(0)
        return output

    # ── Excel (.xlsx) ────────────────────────────────────────────────────

    @staticmethod
    def export_to_excel(title, headers, rows, filename='report.xlsx',
                        summary=None, right_to_left=True):
        """
        Export data to a styled .xlsx workbook.

        Args:
            title:     Sheet / report title
            headers:   list of header strings
            rows:      list of lists (one per data row)
            filename:  download filename
            summary:   optional dict of {label: value} appended below data
            right_to_left: RTL sheet for Arabic

        Returns:
            BytesIO containing the xlsx file.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit

        if right_to_left:
            ws.sheet_view.rightToLeft = True

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4',
                                  fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        money_fmt = '#,##0.000'

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1,
                       end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14, color='4472C4')
        title_cell.alignment = Alignment(horizontal='center')

        # Timestamp
        ts_row = 2
        ws.merge_cells(start_row=ts_row, start_column=1, end_row=ts_row,
                       end_column=len(headers))
        ts_cell = ws.cell(row=ts_row, column=1,
                          value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}")
        ts_cell.alignment = Alignment(horizontal='center')
        ts_cell.font = Font(italic=True, color='888888')

        # Headers
        header_row = 4
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data rows
        for row_idx, row_data in enumerate(rows, header_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                # Auto-format numbers
                if isinstance(value, (int, float)):
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal='right',
                                               vertical='center')

        # Summary section
        if summary:
            summary_row = header_row + len(rows) + 2
            ws.merge_cells(start_row=summary_row, start_column=1,
                           end_row=summary_row, end_column=len(headers))
            ws.cell(row=summary_row, column=1,
                    value='Summary / الإجماليات').font = Font(bold=True, size=12)

            for i, (label, value) in enumerate(summary.items()):
                r = summary_row + 1 + i
                ws.cell(row=r, column=1, value=label).font = Font(bold=True)
                c = ws.cell(row=r, column=2, value=value)
                if isinstance(value, (int, float)):
                    c.number_format = money_fmt

        # Auto-width columns (approximate)
        for col_idx in range(1, len(headers) + 1):
            max_len = max(len(str(headers[col_idx - 1])), 12)
            for row_idx in range(header_row + 1,
                                 header_row + len(rows) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 40))
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = max_len + 2

        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ── PDF (via weasyprint) ─────────────────────────────────────────────

    @staticmethod
    def export_to_pdf(title, headers, rows, filename='report.pdf',
                      summary=None, right_to_left=True):
        """
        Render data as a styled PDF via weasyprint.

        Args:
            title:     Report title
            headers:   list of header strings
            rows:      list of lists
            filename:  download filename
            summary:   optional dict of {label: value}
            right_to_left: RTL layout

        Returns:
            BytesIO containing the PDF file.
        """
        from weasyprint import HTML

        dir_attr = 'dir="rtl"' if right_to_left else ''
        align = 'right' if right_to_left else 'left'

        rows_html = ''
        for row in rows:
            cells = ''.join(
                f'<td style="text-align:{align};padding:6px 8px;border:1px solid #ddd">{cell}</td>'
                for cell in row
            )
            rows_html += f'<tr>{cells}</tr>'

        summary_html = ''
        if summary:
            rows_s = ''.join(
                f'<tr><td style="font-weight:bold;padding:4px 8px">{k}</td>'
                f'<td style="text-align:right;padding:4px 8px">{v}</td></tr>'
                for k, v in summary.items()
            )
            summary_html = f'''
            <h3 style="margin-top:20px;color:#4472C4">Summary / الإجماليات</h3>
            <table style="border-collapse:collapse;width:50%">{rows_s}</table>
            '''

        html_content = f"""<!DOCTYPE html>
<html {dir_attr} lang="ar">
<head>
<meta charset="UTF-8">
<style>
  body {{ font-family: Arial, sans-serif; padding: 20px; {f'direction:rtl;' if right_to_left else ''} }}
  h1 {{ color: #4472C4; border-bottom: 3px solid #4472C4; padding-bottom: 8px; }}
  .meta {{ color: #888; font-size: 0.9em; margin-bottom: 16px; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
  th {{ background: #4472C4; color: #fff; padding: 8px; border: 1px solid #4472C4; }}
  td {{ padding: 6px 8px; border: 1px solid #ddd; }}
  tr:nth-child(even) {{ background: #f5f7fa; }}
  .footer {{ margin-top: 30px; text-align: center; color: #aaa; font-size: 0.8em; }}
</style>
</head>
<body>
  <h1>{title}</h1>
  <div class="meta">Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC</div>
  <table>
    <thead><tr>{''.join(f'<th>{h}</th>' for h in headers)}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  {summary_html}
  <div class="footer">Azad Systems — UAE-Sale ERP</div>
</body>
</html>"""

        pdf_bytes = HTML(string=html_content).write_pdf()
        output = BytesIO(pdf_bytes)
        output.seek(0)
        return output

    # ── Convenience: build rows from dicts ───────────────────────────────

    @staticmethod
    def dicts_to_rows(items, keys):
        """Convert a list of dicts to a list of lists using `keys` order."""
        return [[item.get(k, '') for k in keys] for item in items]

    # ── Legacy CSV helpers (kept for backward compat) ────────────────────

    @staticmethod
    def export_purchases_to_csv(purchases):
        headers = ['الرقم', 'الباقة', 'العميل', 'البريد الإلكتروني',
                   'المبلغ', 'طريقة الدفع', 'الحالة', 'التاريخ']
        data = []
        for p in purchases:
            data.append([
                p.id,
                p.package.name_ar if p.package else 'N/A',
                p.customer_name,
                p.customer_email,
                f'${p.amount_paid}',
                p.payment_method,
                p.payment_status,
                p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else 'N/A'
            ])
        return ExportService.export_to_csv(data, headers, 'purchases.csv')

    @staticmethod
    def export_donations_to_csv(donations):
        headers = ['الرقم', 'المتبرع', 'البريد الإلكتروني', 'المبلغ',
                   'طريقة الدفع', 'الحالة', 'التاريخ']
        data = []
        for d in donations:
            data.append([
                d.id,
                d.donor_name or 'مجهول',
                d.donor_email or 'N/A',
                f'${d.amount_usd}',
                d.payment_method,
                d.status,
                d.created_at.strftime('%Y-%m-%d %H:%M') if d.created_at else 'N/A'
            ])
        return ExportService.export_to_csv(data, headers, 'donations.csv')

    @staticmethod
    def export_cards_to_csv(cards):
        headers = ['الرقم', 'العميل', 'البريد الإلكتروني', 'البطاقة',
                   'النوع', 'المبلغ', 'الحالة', 'التاريخ']
        data = []
        for c in cards:
            data.append([
                c.id,
                c.customer_name,
                c.customer_email,
                c.get_card_display() if hasattr(c, 'get_card_display') else 'N/A',
                c.card_type or 'Unknown',
                f'${c.amount}' if c.amount else 'N/A',
                c.status,
                c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else 'N/A'
            ])
        return ExportService.export_to_csv(data, headers, 'cards.csv')

    # ── Legacy PDF (HTML string) ─────────────────────────────────────────

    @staticmethod
    def generate_pdf_report(title, data, filename='report.pdf'):
        """Legacy: returns an HTML string for browser printing."""
        html = f"""<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
<meta charset="UTF-8"><title>{title}</title>
<style>
body {{ font-family: Arial, sans-serif; direction: rtl; padding: 20px; }}
h1 {{ color: #667eea; border-bottom: 3px solid #667eea; padding-bottom: 10px; }}
.stats {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; margin: 30px 0; }}
.stat-box {{ background: #f8f9fa; padding: 20px; border-radius: 10px; text-align: center; border: 2px solid #667eea; }}
.stat-number {{ font-size: 2rem; color: #667eea; font-weight: bold; }}
.stat-label {{ color: #666; margin-top: 10px; }}
table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
th, td {{ border: 1px solid #ddd; padding: 12px; text-align: right; }}
th {{ background-color: #667eea; color: white; }}
tr:nth-child(even) {{ background-color: #f8f9fa; }}
</style>
</head>
<body>
<h1>{title}</h1>
<p><strong>تاريخ التقرير:</strong> {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}</p>
<div class="stats">{ExportService._generate_stats_html(data.get('stats', {}))}</div>
{ExportService._generate_table_html(data.get('table_data', []), data.get('table_headers', []))}
</body></html>"""
        return html

    @staticmethod
    def _generate_stats_html(stats):
        return ''.join(
            f'<div class="stat-box"><div class="stat-number">{v}</div>'
            f'<div class="stat-label">{k}</div></div>'
            for k, v in stats.items()
        )

    @staticmethod
    def _generate_table_html(data, headers):
        if not data or not headers:
            return ''
        html = '<table><thead><tr>' + ''.join(f'<th>{h}</th>' for h in headers) + '</tr></thead><tbody>'
        for row in data:
            html += '<tr>' + ''.join(f'<td>{c}</td>' for c in row) + '</tr>'
        html += '</tbody></table>'
        return html
