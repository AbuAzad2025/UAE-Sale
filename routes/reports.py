from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from flask import Blueprint, render_template, request, jsonify, make_response
from flask_login import login_required, current_user
from sqlalchemy import func
from extensions import db
from models import Sale, SaleLine, Purchase, Product, Customer, ProductPartner
from utils.decorators import permission_required

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


@reports_bp.route('/')
@login_required
@permission_required('view_reports')
def index():
    return render_template('reports/index.html')


@reports_bp.route('/partners')
@login_required
@permission_required('view_reports')
def partners():  # noqa: C901
    """تقرير الشركاء والمنتجات التابعة للتجار"""
    from models import Payment, Receipt, Supplier

    date_from = request.args.get('date_from', '', type=str)
    date_to = request.args.get('date_to', '', type=str)

    # --- 1. PRODUCT SHARES CALCULATION ---
    # Find products that have partners
    partner_products = Product.query.join(ProductPartner).filter(Product.is_active == True).distinct().all()  # noqa: E712

    partners_data = []
    # Dictionary to aggregate shares per partner: {partner_id: total_share_amount}
    partner_share_totals = {}

    for product in partner_products:
        # Calculate total sales for this product within date range
        sales_query = SaleLine.query.join(Sale).filter(
            SaleLine.product_id == product.id,
            Sale.status == 'confirmed'
        )

        if date_from:
            sales_query = sales_query.filter(func.date(Sale.sale_date) >= date_from)
        if date_to:
            sales_query = sales_query.filter(func.date(Sale.sale_date) <= date_to)

        sales_lines = sales_query.all()

        total_revenue = sum(line.line_total for line in sales_lines)
        total_qty = sum(line.quantity for line in sales_lines)

        # Calculate average unit price
        avg_unit_price = total_revenue / total_qty if total_qty > 0 else 0

        if total_revenue > 0:
            for share in product.partner_shares:
                percentage = Decimal(str(share.percentage))
                partner_amount = total_revenue * (percentage / Decimal('100'))
                partners_data.append({
                    'product_name': product.name,
                    'partner_name': share.partner_customer.name,
                    'percentage': share.percentage,
                    'avg_unit_price': avg_unit_price,
                    'total_qty': total_qty,
                    'total_revenue': total_revenue,
                    'partner_share_amount': partner_amount
                })

                # Aggregate for summary
                p_id = share.partner_customer.id
                partner_share_totals[p_id] = partner_share_totals.get(p_id, Decimal('0')) + partner_amount

    # Find products linked to a merchant
    merchant_products = Product.query.filter(
        Product.merchant_customer_id.isnot(None),
        Product.is_active.is_(True)
    ).all()

    merchants_data = []
    # Dictionary to aggregate shares per merchant: {merchant_id: total_share_amount}
    merchant_share_totals = {}

    for product in merchant_products:
        sales_query = SaleLine.query.join(Sale).filter(
            SaleLine.product_id == product.id,
            Sale.status == 'confirmed'
        )

        if date_from:
            sales_query = sales_query.filter(func.date(Sale.sale_date) >= date_from)
        if date_to:
            sales_query = sales_query.filter(func.date(Sale.sale_date) <= date_to)

        sales_lines = sales_query.all()

        total_revenue = sum(line.line_total for line in sales_lines)
        total_qty = sum(line.quantity for line in sales_lines)

        # Calculate average unit price
        avg_unit_price = total_revenue / total_qty if total_qty > 0 else 0

        if total_revenue > 0:
            merchant_percentage = float(product.merchant_share or 100)
            merchant_amount = total_revenue * (Decimal(merchant_percentage) / 100)

            merchants_data.append({
                'product_name': product.name,
                'merchant_name': product.merchant_customer.name,
                'percentage': merchant_percentage,
                'avg_unit_price': avg_unit_price,
                'total_qty': total_qty,
                'total_revenue': total_revenue,
                'merchant_share_amount': merchant_amount
            })

            # Aggregate for summary
            m_id = product.merchant_customer.id
            merchant_share_totals[m_id] = merchant_share_totals.get(m_id, Decimal('0')) + merchant_amount

    # --- 2. FINANCIAL SUMMARIES (Partners & Merchants) ---
    # Helper to get payments/receipts
    def get_financials(customer_type, share_totals_dict):
        customers = Customer.query.filter_by(customer_type=customer_type).all()
        summary_list = []

        for cust in customers:
            # Paid TO Customer (Outgoing Payments)
            paid_query = db.session.query(func.sum(Payment.amount_aed)).filter(
                Payment.customer_id == cust.id,
                Payment.direction == 'outgoing'
            )
            # Received FROM Customer (Receipts OR Incoming Payments)
            # 1. Receipts
            receipts_query = db.session.query(func.sum(Receipt.amount_aed)).filter(
                Receipt.customer_id == cust.id
            )
            # 2. Incoming Payments (Refunds/etc)
            payment_in_query = db.session.query(func.sum(Payment.amount_aed)).filter(
                Payment.customer_id == cust.id,
                Payment.direction == 'incoming'
            )

            if date_from:
                paid_query = paid_query.filter(func.date(Payment.payment_date) >= date_from)
                receipts_query = receipts_query.filter(func.date(Receipt.receipt_date) >= date_from)
                payment_in_query = payment_in_query.filter(func.date(Payment.payment_date) >= date_from)
            if date_to:
                paid_query = paid_query.filter(func.date(Payment.payment_date) <= date_to)
                receipts_query = receipts_query.filter(func.date(Receipt.receipt_date) <= date_to)
                payment_in_query = payment_in_query.filter(func.date(Payment.payment_date) <= date_to)

            total_paid_to = paid_query.scalar() or Decimal('0')
            total_receipts = receipts_query.scalar() or Decimal('0')
            total_payment_in = payment_in_query.scalar() or Decimal('0')
            total_received_from = total_receipts + total_payment_in

            total_share = share_totals_dict.get(cust.id, Decimal('0'))

            # For Partner/Merchant:
            # Balance (Net) = (Total Share + Total Received From) - Total Paid To
            # Assuming 'Share' is money they earned (credit to them).
            # 'Received From' is money they gave us (credit to them, or debt repayment?).
            # Usually: Balance = (Earnings + Deposits) - Withdrawals
            net_balance = (total_share + total_received_from) - total_paid_to

            # Only add if there's any activity
            if total_share > 0 or total_paid_to > 0 or total_received_from > 0:
                summary_list.append({
                    'name': cust.name,
                    'total_share': total_share,
                    'paid_to': total_paid_to,
                    'received_from': total_received_from,
                    'net_balance': net_balance
                })
        return summary_list

    partners_summary = get_financials('partner', partner_share_totals)
    merchants_summary = get_financials('merchant', merchant_share_totals)

    # --- 3. SUPPLIERS SUMMARY ---
    suppliers = Supplier.query.all()
    suppliers_summary = []

    for sup in suppliers:
        # Total Purchases
        purchases_query = db.session.query(func.sum(Purchase.amount_aed)).filter(
            Purchase.supplier_id == sup.id,
            Purchase.status == 'confirmed'
        )
        # Paid TO Supplier (Outgoing)
        paid_query = db.session.query(func.sum(Payment.amount_aed)).filter(
            Payment.supplier_id == sup.id,
            Payment.direction == 'outgoing'
        )
        # Received FROM Supplier (Incoming - Refunds)
        received_query = db.session.query(func.sum(Payment.amount_aed)).filter(
            Payment.supplier_id == sup.id,
            Payment.direction == 'incoming'
        )

        if date_from:
            purchases_query = purchases_query.filter(func.date(Purchase.purchase_date) >= date_from)
            paid_query = paid_query.filter(func.date(Payment.payment_date) >= date_from)
            received_query = received_query.filter(func.date(Payment.payment_date) >= date_from)
        if date_to:
            purchases_query = purchases_query.filter(func.date(Purchase.purchase_date) <= date_to)
            paid_query = paid_query.filter(func.date(Payment.payment_date) <= date_to)
            received_query = received_query.filter(func.date(Payment.payment_date) <= date_to)

        total_purchases = purchases_query.scalar() or Decimal('0')
        total_paid_to = paid_query.scalar() or Decimal('0')
        total_refunds = received_query.scalar() or Decimal('0')

        # Balance = Purchases - (Paid - Refunds)
        # Or: Purchases - Net Paid
        net_paid = total_paid_to - total_refunds
        balance_due = total_purchases - net_paid

        if total_purchases > 0 or total_paid_to > 0 or total_refunds > 0:
            suppliers_summary.append({
                'name': sup.name,
                'total_purchases': total_purchases,
                'paid_to': total_paid_to,
                'received_from': total_refunds,
                'balance_due': balance_due
            })

    return render_template('reports/partners.html',
                           partners_data=partners_data,
                           merchants_data=merchants_data,
                           partners_summary=partners_summary,
                           merchants_summary=merchants_summary,
                           suppliers_summary=suppliers_summary)


@reports_bp.route('/sales')
@login_required
@permission_required('view_reports')
def sales():
    date_from = request.args.get('date_from', '', type=str)
    date_to = request.args.get('date_to', '', type=str)
    customer_id = request.args.get('customer', type=int)
    seller_id = request.args.get('seller', type=int)

    query = Sale.query.filter_by(status='confirmed')

    if date_from:
        query = query.filter(func.date(Sale.sale_date) >= date_from)

    if date_to:
        query = query.filter(func.date(Sale.sale_date) <= date_to)

    if customer_id:
        query = query.filter_by(customer_id=customer_id)

    if seller_id:
        query = query.filter_by(seller_id=seller_id)
    elif current_user.is_seller():
        query = query.filter_by(seller_id=current_user.id)

    sales_list = query.order_by(Sale.sale_date.desc()).all()

    total_sales = Decimal('0')
    total_paid = Decimal('0')
    total_due = Decimal('0')

    for sale in sales_list:
        total_sales += (sale.amount_aed or Decimal('0'))
        total_paid += (sale.paid_amount_aed or Decimal('0'))
        total_due += ((sale.amount_aed or Decimal('0')) - (sale.paid_amount_aed or Decimal('0')))

    total_profit = Decimal('0')
    if current_user.can_see_costs():
        for sale in sales_list:
            total_profit += (sale.get_profit() or Decimal('0'))

    summary = {
        'sales_count': len(sales_list),
        'total_sales_aed': float(total_sales),
        'total_paid_aed': float(total_paid),
        'total_pending_aed': float(total_due),
        'total_profit': float(total_profit) if current_user.can_see_costs() else None
    }

    return render_template('reports/sales.html',
                           sales=sales_list,
                           summary=summary)


@reports_bp.route('/purchases')
@login_required
@permission_required('view_reports')
def purchases():  # noqa: C901
    if current_user.is_seller():
        return render_template('errors/403.html'), 403

    date_from = request.args.get('start_date', '', type=str)
    date_to = request.args.get('end_date', '', type=str)
    supplier_id = request.args.get('supplier_id', type=int)

    query = Purchase.query.filter_by(status='confirmed')

    if date_from:
        query = query.filter(func.date(Purchase.purchase_date) >= date_from)

    if date_to:
        query = query.filter(func.date(Purchase.purchase_date) <= date_to)

    if supplier_id:
        query = query.filter_by(supplier_id=supplier_id)

    purchases_list = query.order_by(Purchase.purchase_date.desc()).all()

    total_amount = Decimal('0')

    # Calculate total purchases amount
    for p in purchases_list:
        amount = p.amount_aed or Decimal('0')
        total_amount += amount
        # Add dummy attributes for template compatibility if not present
        if not hasattr(p, 'paid_amount'):
            p.paid_amount = Decimal('0')
        if not hasattr(p, 'balance_due'):
            p.balance_due = amount

    # Calculate total paid from Payments table
    from models import Payment
    payment_query = Payment.query.filter(
        Payment.direction == 'outgoing',
        Payment.supplier_id is not None
    )

    if date_from:
        payment_query = payment_query.filter(func.date(Payment.payment_date) >= date_from)
    if date_to:
        payment_query = payment_query.filter(func.date(Payment.payment_date) <= date_to)
    if supplier_id:
        payment_query = payment_query.filter_by(supplier_id=supplier_id)

    payments_list = payment_query.all()
    total_paid = sum((p.amount_aed or Decimal('0') for p in payments_list), Decimal('0'))

    # Total due is rough estimate: Purchases - Payments
    # Note: This doesn't account for opening balance
    total_due = total_amount - total_paid

    stats = {
        'total_purchases': len(purchases_list),
        'total_amount': float(total_amount),
        'total_paid': float(total_paid),
        'total_due': float(total_due)
    }

    # Get suppliers for filter
    from models import Supplier
    suppliers = Supplier.query.filter_by(is_active=True).all()

    return render_template('reports/purchases.html',
                           purchases=purchases_list,
                           stats=stats,
                           suppliers=suppliers,
                           start_date=date_from,
                           end_date=date_to,
                           supplier_id=supplier_id)


@reports_bp.route('/receivables')
@login_required
@permission_required('view_reports')
def receivables():
    now = datetime.now(timezone.utc)

    all_sales = Sale.query.filter(
        Sale.status == 'confirmed'
    ).all()

    all_sales = [sale for sale in all_sales if (sale.amount_aed or Decimal('0')) > (sale.paid_amount_aed or Decimal('0'))]

    aging_data = {
        'current': {'sales': [], 'total': Decimal('0')},
        'days_30': {'sales': [], 'total': Decimal('0')},
        'days_60': {'sales': [], 'total': Decimal('0')},
        'days_90': {'sales': [], 'total': Decimal('0')},
        'over_90': {'sales': [], 'total': Decimal('0')},
    }

    for sale in all_sales:
        sale_date = sale.sale_date
        if sale_date.tzinfo is None:
            sale_date = sale_date.replace(tzinfo=timezone.utc)
        days_old = (now - sale_date).days
        balance = (sale.amount_aed or Decimal('0')) - (sale.paid_amount_aed or Decimal('0'))

        sale.days_old = days_old
        sale.calculated_balance = balance

        if days_old <= 30:
            aging_data['current']['sales'].append(sale)
            aging_data['current']['total'] += balance
        elif days_old <= 60:
            aging_data['days_30']['sales'].append(sale)
            aging_data['days_30']['total'] += balance
        elif days_old <= 90:
            aging_data['days_60']['sales'].append(sale)
            aging_data['days_60']['total'] += balance
        elif days_old <= 120:
            aging_data['days_90']['sales'].append(sale)
            aging_data['days_90']['total'] += balance
        else:
            aging_data['over_90']['sales'].append(sale)
            aging_data['over_90']['total'] += balance

    total_receivables = sum(data['total'] for data in aging_data.values())

    summary = {
        'total_receivables': float(total_receivables),
        'current': float(aging_data['current']['total']),
        'days_30': float(aging_data['days_30']['total']),
        'days_60': float(aging_data['days_60']['total']),
        'days_90': float(aging_data['days_90']['total']),
        'over_90': float(aging_data['over_90']['total']),
    }

    return render_template('reports/receivables.html',
                           aging_data=aging_data,
                           summary=summary)


@reports_bp.route('/inventory')
@login_required
@permission_required('view_reports')
def inventory():
    category_id = request.args.get('category', type=int)

    query = Product.query.filter_by(is_active=True)

    if category_id:
        query = query.filter_by(category_id=category_id)

    products = query.order_by(Product.name).all()

    total_value = Decimal('0')
    total_items = Decimal('0')

    for p in products:
        total_items += (p.current_stock or Decimal('0'))
        if current_user.can_see_costs():
            total_value += (p.current_stock or Decimal('0')) * (p.cost_price or Decimal('0'))

    summary = {
        'products_count': len(products),
        'total_items': float(total_items),
        'total_value': float(total_value) if current_user.can_see_costs() else None
    }

    return render_template('reports/inventory.html',
                           products=products,
                           summary=summary)


@reports_bp.route('/api/entity-search')
@login_required
@permission_required('view_reports')
def api_entity_search():
    from models import Supplier
    query = request.args.get('q', '').strip()
    entity_type = request.args.get('type', 'supplier')

    results = []

    if entity_type == 'supplier':
        suppliers = Supplier.query.filter(
            db.or_(
                Supplier.name.ilike(f'%{query}%'),
                Supplier.phone.ilike(f'%{query}%')
            )
        ).limit(10).all()
        for s in suppliers:
            results.append({
                'id': s.id,
                'name': s.name,
                'phone': s.phone,
                'type': 'supplier'
            })

    else:  # customer, partner, merchant
        q_filter = Customer.query.filter(
            db.or_(
                Customer.name.ilike(f'%{query}%'),
                Customer.phone.ilike(f'%{query}%')
            )
        )

        if entity_type == 'partner':
            q_filter = q_filter.filter_by(customer_type='partner')
        elif entity_type == 'merchant':
            q_filter = q_filter.filter_by(customer_type='merchant')
        # If 'customer', we search all or just regular? Let's search all if type is generic, or filter if specific.
        # User dropdown will likely send specific types.

        customers = q_filter.limit(10).all()
        for c in customers:
            results.append({
                'id': c.id,
                'name': c.name,
                'phone': c.phone,
                'type': c.customer_type
            })

    return jsonify(results)


@reports_bp.route('/entity_report_fragment/<type>/<id>')
@login_required
@permission_required('view_reports')
def entity_report_fragment(type, id):  # noqa: C901
    try:
        from models import Receipt, Payment, PurchaseLine, Supplier

        context = {
            'entity': None,
            'type_label': '',
            'balance': 0,
            'balance_label': '',
            'products': [],
            'invoices': [],
            'transactions': []
        }

        if type == 'supplier':
            entity = db.get_or_404(Supplier, id)
            context['entity'] = entity
            context['type_label'] = 'مورد'

            # Balance
            context['balance'] = entity.get_balance_aed()
            context['balance_label'] = 'مستحق للمورد'

            # Products (Purchased)
            p_lines = db.session.query(
                Product.name,
                func.sum(PurchaseLine.quantity).label('qty'),
                func.sum(PurchaseLine.line_total).label('total'),
                func.max(Purchase.purchase_date).label('last_date')
            ).join(Purchase).join(Product).filter(
                Purchase.supplier_id == id,
                Purchase.status == 'confirmed'
            ).group_by(Product.name).all()

            context['products'] = [{
                'name': p.name,
                'quantity': p.qty,
                'total': p.total,
                'last_date': p.last_date.strftime('%Y-%m-%d') if p.last_date else '-'
            } for p in p_lines]

            # Invoices (Purchases)
            purchases = Purchase.query.filter_by(supplier_id=id).order_by(Purchase.purchase_date.desc()).all()
            context['invoices'] = [{
                'number': p.purchase_number,
                'date': p.purchase_date.strftime('%Y-%m-%d'),
                'status': p.status,
                'amount': p.amount_aed or 0,
                'paid': p.get_paid_amount(),
                'balance': (p.amount_aed or 0) - p.get_paid_amount()
            } for p in purchases]

            # Transactions (Payments TO Supplier)
            payments = Payment.query.filter_by(supplier_id=id).order_by(Payment.payment_date.desc()).all()
            context['transactions'] = [{
                'number': p.payment_number,
                'type': 'out',  # Payment out
                'date': p.payment_date.strftime('%Y-%m-%d'),
                'amount': p.amount_aed,
                'method': p.payment_method,
                'notes': p.notes or '-'
            } for p in payments]

        else:  # Customer/Partner/Merchant
            entity = db.get_or_404(Customer, id)
            context['entity'] = entity
            context['type_label'] = {
                'partner': 'شريك',
                'merchant': 'تاجر',
                'regular': 'زبون',
                'vip': 'VIP'
            }.get(entity.customer_type, 'زبون')

            # Balance calculation (Receivables/Payables)
            # Sales (He took goods) + Payments Out (He took money) - Receipts (He gave money)
            total_sales = db.session.query(func.sum(Sale.amount_aed)).filter(Sale.customer_id == id, Sale.status == 'confirmed').scalar() or 0
            total_receipts = db.session.query(func.sum(Receipt.amount_aed)).filter(Receipt.customer_id == id).scalar() or 0
            # Payments made TO customer (e.g. returns/share/drawings)
            total_payments_to = db.session.query(func.sum(Payment.amount_aed)).filter(Payment.customer_id == id, Payment.direction == 'outgoing').scalar() or 0

            context['balance'] = (total_sales + total_payments_to) - total_receipts  # Positive means they owe us
            context['balance_label'] = 'مستحق لنا'
            if context['balance'] < 0:
                context['balance'] = abs(context['balance'])
                context['balance_label'] = 'مستحق للعميل'

            # Products (Sold) - Products the customer BOUGHT
            s_lines = db.session.query(
                Product.name,
                func.sum(SaleLine.quantity).label('qty'),
                func.sum(SaleLine.line_total).label('total'),
                func.max(Sale.sale_date).label('last_date')
            ).join(Sale).join(Product).filter(
                Sale.customer_id == id,
                Sale.status == 'confirmed'
            ).group_by(Product.name).all()

            context['products'] = [{
                'name': p.name,
                'quantity': p.qty,
                'total': p.total,
                'last_date': p.last_date.strftime('%Y-%m-%d') if p.last_date else '-'
            } for p in s_lines]

            # IF PARTNER: Fetch products they have a share in (Products they EARN from)
            if entity.customer_type == 'partner':
                shared_products_query = db.session.query(
                    Product.name,
                    ProductPartner.percentage,
                    func.sum(SaleLine.quantity).label('qty'),
                    func.sum(SaleLine.line_total).label('total_sales'),
                    func.max(Sale.sale_date).label('last_date')
                ).join(ProductPartner, Product.id == ProductPartner.product_id)\
                 .join(SaleLine, SaleLine.product_id == Product.id)\
                 .join(Sale, Sale.id == SaleLine.sale_id)\
                 .filter(
                     ProductPartner.partner_customer_id == id,
                     Sale.status == 'confirmed'
                 ).group_by(Product.name, ProductPartner.percentage).all()

                for sp in shared_products_query:
                    share_amount = sp.total_sales * (sp.percentage / 100)
                    context['products'].append({
                        'name': f"{sp.name} (Share: {sp.percentage}%)",
                        'quantity': sp.qty,
                        'total': share_amount,
                        'last_date': sp.last_date.strftime('%Y-%m-%d') if sp.last_date else '-'
                    })

            # IF MERCHANT: Fetch products they own (Products they EARN from)
            if entity.customer_type == 'merchant':
                merchant_products_query = db.session.query(
                    Product.name,
                    Product.merchant_share,
                    func.sum(SaleLine.quantity).label('qty'),
                    func.sum(SaleLine.line_total).label('total_sales'),
                    func.max(Sale.sale_date).label('last_date')
                ).join(SaleLine, SaleLine.product_id == Product.id)\
                 .join(Sale, Sale.id == SaleLine.sale_id)\
                 .filter(
                     Product.merchant_customer_id == id,
                     Sale.status == 'confirmed'
                 ).group_by(Product.name, Product.merchant_share).all()

                for mp in merchant_products_query:
                    share_pct = mp.merchant_share or 100
                    share_amount = mp.total_sales * (share_pct / 100)
                    context['products'].append({
                        'name': f"{mp.name} (Merchant: {share_pct}%)",
                        'quantity': mp.qty,
                        'total': share_amount,
                        'last_date': mp.last_date.strftime('%Y-%m-%d') if mp.last_date else '-'
                    })

            # Invoices (Sales)
            sales = Sale.query.filter_by(customer_id=id).order_by(Sale.sale_date.desc()).all()
            context['invoices'] = [{
                'number': s.sale_number,
                'date': s.sale_date.strftime('%Y-%m-%d'),
                'status': s.status,
                'amount': s.amount_aed or 0,
                'paid': s.paid_amount_aed or 0,
                'balance': (s.amount_aed or 0) - (s.paid_amount_aed or 0)
            } for s in sales]

            # Transactions (Receipts + Payments)
            receipts = Receipt.query.filter_by(customer_id=id).all()
            payments_out = Payment.query.filter_by(customer_id=id, direction='outgoing').all()

            all_trans = []
            for r in receipts:
                all_trans.append({
                    'number': r.receipt_number,
                    'type': 'in',  # Money In
                    'date': r.receipt_date,
                    'amount': r.amount_aed,
                    'method': r.payment_method,
                    'notes': 'قبض'
                })
            for p in payments_out:
                all_trans.append({
                    'number': p.payment_number,
                    'type': 'out',  # Money Out
                    'date': p.payment_date,
                    'amount': p.amount_aed,
                    'method': p.payment_method,
                    'notes': p.notes or 'دفع'
                })

            all_trans.sort(key=lambda x: x['date'], reverse=True)
            for t in all_trans:
                t['date'] = t['date'].strftime('%Y-%m-%d')

            context['transactions'] = all_trans

        return render_template('reports/partials/entity_report.html', **context)

    except Exception as e:
        return render_template('reports/partials/entity_report.html', error=str(e))


@reports_bp.route('/top-selling')
@login_required
@permission_required('view_reports')
def top_selling():
    date_from = request.args.get('date_from', '', type=str)
    date_to = request.args.get('date_to', '', type=str)
    limit = request.args.get('limit', 20, type=int)

    query = db.session.query(
        Product.id,
        Product.name,
        func.sum(SaleLine.quantity).label('total_quantity'),
        func.sum(SaleLine.line_total).label('total_sales')
    ).join(
        SaleLine, Product.id == SaleLine.product_id
    ).join(
        Sale, SaleLine.sale_id == Sale.id
    ).filter(
        Sale.status == 'confirmed'
    )

    if date_from:
        query = query.filter(func.date(Sale.sale_date) >= date_from)

    if date_to:
        query = query.filter(func.date(Sale.sale_date) <= date_to)

    products = query.group_by(
        Product.id, Product.name
    ).order_by(
        func.sum(SaleLine.quantity).desc()
    ).limit(limit).all()

    return render_template('reports/top_selling.html', products=products)


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 2: NEW REPORTS
# ═══════════════════════════════════════════════════════════════════════════════


@reports_bp.route('/inventory-valuation')
@login_required
@permission_required('view_reports')
def inventory_valuation():
    """
    تقرير تقييم المخزون — Inventory Valuation
    Per warehouse/category, qty × cost, totals.
    """
    from models import Warehouse, ProductCategory

    warehouse_id = request.args.get('warehouse_id', type=int)
    category_id = request.args.get('category_id', type=int)

    # Base query: active products with stock > 0
    products_query = Product.query.filter(
        Product.is_active.is_(True),
        Product.current_stock > 0
    )

    if category_id:
        products_query = products_query.filter_by(category_id=category_id)

    products = products_query.order_by(Product.name).all()

    # Group by category
    categories_data = {}
    total_value = Decimal('0')
    total_qty = Decimal('0')

    for product in products:
        qty = product.current_stock or Decimal('0')
        cost = product.cost_price or Decimal('0')
        value = qty * cost
        total_value += value
        total_qty += qty

        cat_name = product.category.name if product.category else 'بدون فئة'
        cat_name_ar = product.category.name_ar if product.category and product.category.name_ar else cat_name

        if cat_name not in categories_data:
            categories_data[cat_name] = {
                'name': cat_name,
                'name_ar': cat_name_ar,
                'products': [],
                'total_qty': Decimal('0'),
                'total_value': Decimal('0'),
                'product_count': 0,
            }

        categories_data[cat_name]['products'].append({
            'product': product,
            'qty': qty,
            'cost': cost,
            'value': value,
        })
        categories_data[cat_name]['total_qty'] += qty
        categories_data[cat_name]['total_value'] += value
        categories_data[cat_name]['product_count'] += 1

    warehouses = Warehouse.query.filter_by(is_active=True).all()
    categories = ProductCategory.query.filter_by(is_active=True).all()

    return render_template('reports/inventory_valuation.html',
                           categories_data=categories_data,
                           warehouses=warehouses,
                           categories_list=categories,
                           total_value=total_value,
                           total_qty=total_qty,
                           total_products=products.count(),
                           selected_warehouse=warehouse_id,
                           selected_category=category_id)


@reports_bp.route('/ap-aging')
@login_required
@permission_required('view_reports')
def ap_aging():  # noqa: C901
    """
    تقرير أعمار الذمم الدائنة — Accounts Payable Aging
    Supplier balances bucketed 0–30/31–60/61–90/90+ days.
    """
    from models import Supplier, Purchase, Payment
    now = datetime.now(timezone.utc)

    suppliers = Supplier.query.filter_by(is_active=True).all()

    supplier_aging = []
    total_payable = Decimal('0')

    aging_buckets = {
        'current': Decimal('0'),    # 0-30 days
        'days_31_60': Decimal('0'),  # 31-60 days
        'days_61_90': Decimal('0'),  # 61-90 days
        'over_90': Decimal('0'),    # 90+ days
    }

    for supplier in suppliers:
        # Get all confirmed purchases for this supplier
        purchases = Purchase.query.filter(
            Purchase.supplier_id == supplier.id,
            Purchase.status == 'confirmed'
        ).all()

        # Get all confirmed outgoing payments for this supplier
        payments = Payment.query.filter(
            Payment.supplier_id == supplier.id,
            Payment.direction == 'outgoing',
            Payment.payment_confirmed.is_(True)
        ).all()

        total_purchases = sum((p.amount_aed or Decimal('0') for p in purchases), Decimal('0'))
        total_paid = sum((p.amount_aed or Decimal('0') for p in payments), Decimal('0'))
        balance = total_purchases - total_paid

        if balance <= 0:
            continue

        # Age the balance using the oldest unpaid purchase
        oldest_purchase_date = None
        for p in purchases:
            if p.purchase_date:
                if oldest_purchase_date is None or p.purchase_date < oldest_purchase_date:
                    oldest_purchase_date = p.purchase_date

        days_old = 0
        if oldest_purchase_date:
            if oldest_purchase_date.tzinfo is None:
                oldest_purchase_date = oldest_purchase_date.replace(tzinfo=timezone.utc)
            days_old = (now - oldest_purchase_date).days

        bucket = 'current'
        if days_old > 90:
            bucket = 'over_90'
        elif days_old > 60:
            bucket = 'days_61_90'
        elif days_old > 30:
            bucket = 'days_31_60'

        aging_buckets[bucket] += balance
        total_payable += balance

        supplier_aging.append({
            'supplier': supplier,
            'total_purchases': total_purchases,
            'total_paid': total_paid,
            'balance': balance,
            'days_old': days_old,
            'bucket': bucket,
            'payment_terms': supplier.payment_terms_days or 30,
        })

    # Sort by oldest first
    supplier_aging.sort(key=lambda x: x['days_old'], reverse=True)

    return render_template('reports/ap_aging.html',
                           supplier_aging=supplier_aging,
                           aging_buckets=aging_buckets,
                           total_payable=total_payable)


@reports_bp.route('/cash-flow')
@login_required
@permission_required('view_reports')
def cash_flow():
    """
    قائمة التدفقات النقدية — Cash Flow Statement
    Operating / Investing / Financing from GL data.
    """
    from models import GLAccount, GLJournalLine, GLJournalEntry

    date_from = request.args.get('date_from', '', type=str)
    date_to = request.args.get('date_to', '', type=str)

    # Build base query for GL lines within date range
    line_query = db.session.query(
        GLJournalLine.account_id,
        GLJournalLine.debit,
        GLJournalLine.credit,
        GLJournalLine.amount_aed,
        GLJournalEntry.entry_date,
        GLAccount.type.label('account_type'),
        GLAccount.code.label('account_code'),
        GLAccount.name.label('account_name'),
        GLAccount.name_ar.label('account_name_ar'),
    ).join(
        GLJournalEntry, GLJournalLine.entry_id == GLJournalEntry.id
    ).join(
        GLAccount, GLJournalLine.account_id == GLAccount.id
    ).filter(
        GLJournalEntry.is_posted.is_(True),
        GLJournalEntry.is_reversed.is_(False)
    )

    if date_from:
        line_query = line_query.filter(func.date(GLJournalEntry.entry_date) >= date_from)
    if date_to:
        line_query = line_query.filter(func.date(GLJournalEntry.entry_date) <= date_to)

    lines = line_query.all()

    # Classify into cash flow categories
    # Operating: revenue (4xxx) and expense (5xxx-6xxx) accounts
    # Investing: asset accounts (1xxx) — fixed assets, equipment
    # Financing: liability (2xxx) and equity (3xxx) accounts
    operating_items = {}
    investing_items = {}
    financing_items = {}

    for line in lines:
        code = line.account_code or ''
        acct_type = line.account_type
        name = line.account_name_ar or line.account_name
        key = f"{code} - {name}"

        # Net amount: debit increases assets/expenses, credit increases liabilities/equity/revenue
        net = (line.debit or Decimal('0')) - (line.credit or Decimal('0'))

        if acct_type == 'revenue':
            # Revenue: credit is positive cash flow
            net = -net  # Flip so credit = positive
            operating_items[key] = operating_items.get(key, Decimal('0')) + net
        elif acct_type == 'expense':
            # Expenses: debit is negative cash flow
            operating_items[key] = operating_items.get(key, Decimal('0')) + net
        elif acct_type == 'asset' and code.startswith(('1',)):
            # Fixed assets (investing)
            investing_items[key] = investing_items.get(key, Decimal('0')) + net
        elif acct_type in ('liability', 'equity'):
            # Financing
            financing_items[key] = financing_items.get(key, Decimal('0')) - net

    operating_total = sum(operating_items.values(), Decimal('0'))
    investing_total = sum(investing_items.values(), Decimal('0'))
    financing_total = sum(financing_items.values(), Decimal('0'))
    net_change = operating_total + investing_total + financing_total

    return render_template('reports/cash_flow.html',
                           operating_items=operating_items,
                           investing_items=investing_items,
                           financing_items=financing_items,
                           operating_total=operating_total,
                           investing_total=investing_total,
                           financing_total=financing_total,
                           net_change=net_change,
                           date_from=date_from,
                           date_to=date_to)


@reports_bp.route('/vat-report')
@login_required
@permission_required('view_reports')
def vat_report():
    """
    تقرير ضريبة القيمة المضافة — VAT Report (UAE 5%)
    Taxable sales/purchases, output/input VAT, net payable.
    """
    _ = Decimal('0.05')  # UAE 5%

    date_from = request.args.get('date_from', '', type=str)
    date_to = request.args.get('date_to', '', type=str)

    # --- Taxable Sales (Output VAT) ---
    sales_query = Sale.query.filter_by(status='confirmed')
    if date_from:
        sales_query = sales_query.filter(func.date(Sale.sale_date) >= date_from)
    if date_to:
        sales_query = sales_query.filter(func.date(Sale.sale_date) <= date_to)

    sales = sales_query.all()

    taxable_sales = []
    total_taxable_sales = Decimal('0')
    total_output_vat = Decimal('0')
    total_exempt_sales = Decimal('0')

    for sale in sales:
        amount = sale.amount_aed or Decimal('0')
        tax_rate = sale.tax_rate or Decimal('0')
        tax_amount = sale.tax_amount or Decimal('0')

        if tax_rate > 0:
            taxable_sales.append({
                'number': sale.sale_number,
                'date': sale.sale_date,
                'customer': sale.customer.name if sale.customer else 'عميل نقدي',
                'amount': amount,
                'tax_rate': tax_rate,
                'tax_amount': tax_amount,
            })
            total_taxable_sales += amount
            total_output_vat += tax_amount
        else:
            total_exempt_sales += amount

    # --- Taxable Purchases (Input VAT) ---
    purchases_query = Purchase.query.filter_by(status='confirmed')
    if date_from:
        purchases_query = purchases_query.filter(func.date(Purchase.purchase_date) >= date_from)
    if date_to:
        purchases_query = purchases_query.filter(func.date(Purchase.purchase_date) <= date_to)

    purchases = purchases_query.all()

    taxable_purchases = []
    total_taxable_purchases = Decimal('0')
    total_input_vat = Decimal('0')
    total_exempt_purchases = Decimal('0')

    for purchase in purchases:
        amount = purchase.amount_aed or Decimal('0')
        tax_rate = purchase.tax_rate or Decimal('0')
        tax_amount = purchase.tax_amount or Decimal('0')

        if tax_rate > 0:
            taxable_purchases.append({
                'number': purchase.purchase_number,
                'date': purchase.purchase_date,
                'supplier': purchase.supplier.name if purchase.supplier else purchase.supplier_name,
                'amount': amount,
                'tax_rate': tax_rate,
                'tax_amount': tax_amount,
            })
            total_taxable_purchases += amount
            total_input_vat += tax_amount
        else:
            total_exempt_purchases += amount

    # Net VAT payable = Output VAT - Input VAT
    net_vat = total_output_vat - total_input_vat

    summary = {
        'total_taxable_sales': total_taxable_sales,
        'total_output_vat': total_output_vat,
        'total_exempt_sales': total_exempt_sales,
        'total_taxable_purchases': total_taxable_purchases,
        'total_input_vat': total_input_vat,
        'total_exempt_purchases': total_exempt_purchases,
        'net_vat': net_vat,
        'vat_payable': net_vat if net_vat > 0 else Decimal('0'),
        'vat_refund': abs(net_vat) if net_vat < 0 else Decimal('0'),
    }

    return render_template('reports/vat_report.html',
                           taxable_sales=taxable_sales,
                           taxable_purchases=taxable_purchases,
                           summary=summary,
                           date_from=date_from,
                           date_to=date_to)


# ═══════════════════════════════════════════════════════════════════════════════
# PHASE 3: EXPORT ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════


def _send_export(export_func, filename, format_type):
    """Helper: generate export and return as downloadable file."""
    output = export_func()
    mime = {
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'pdf': 'application/pdf',
        'csv': 'text/csv; charset=utf-8',
    }.get(format_type, 'application/octet-stream')
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = mime
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@reports_bp.route('/inventory-valuation/export')
@login_required
@permission_required('view_reports')
def export_inventory_valuation():
    from services.export_service import ExportService
    fmt = request.args.get('format', 'xlsx')

    products = Product.query.filter(Product.is_active == True, Product.current_stock > 0).order_by(Product.name).all()  # noqa: E712
    headers = ['المنتج', 'الكود', 'الفئة', 'الكمية', 'سعر التكلفة', 'القيمة']
    rows = []
    total_value = Decimal('0')
    total_qty = Decimal('0')
    for p in products:
        qty = p.current_stock or Decimal('0')
        cost = p.cost_price or Decimal('0')
        val = qty * cost
        total_value += val
        total_qty += qty
        rows.append([p.name, p.sku or p.part_number or '',
                     p.category.name if p.category else '',
                     float(qty), float(cost), float(val)])
    summary = {'إجمالي الكمية': float(total_qty), 'إجمالي القيمة': float(total_value)}

    def build():
        if fmt == 'pdf':
            return ExportService.export_to_pdf('تقييم المخزون — Inventory Valuation',
                                               headers, rows, summary=summary)
        elif fmt == 'csv':
            all_rows = rows + [['', '', 'المجموع', float(total_qty), '', float(total_value)]]
            return ExportService.export_to_csv(all_rows, headers, 'inventory_valuation.csv')
        return ExportService.export_to_excel('تقييم المخزون — Inventory Valuation',
                                             headers, rows, summary=summary)

    ext = {'xlsx': 'xlsx', 'pdf': 'pdf', 'csv': 'csv'}.get(fmt, 'xlsx')
    return _send_export(build, f'inventory_valuation.{ext}', ext)


@reports_bp.route('/ap-aging/export')
@login_required
@permission_required('view_reports')
def export_ap_aging():
    from services.export_service import ExportService
    from models import Supplier, Purchase, Payment
    from datetime import timezone as tz
    fmt = request.args.get('format', 'xlsx')
    now = datetime.now(tz)

    suppliers = Supplier.query.filter_by(is_active=True).all()
    headers = ['المورد', 'إجمالي المشتريات', 'المدفوع', 'الرصيد المستحق', 'الأيام', 'شروط الدفع']
    rows = []
    total_payable = Decimal('0')
    for supplier in suppliers:
        purchases = Purchase.query.filter(Purchase.supplier_id == supplier.id, Purchase.status == 'confirmed').all()
        payments = Payment.query.filter(
            Payment.supplier_id == supplier.id,
            Payment.direction == 'outgoing',
            Payment.payment_confirmed == True,  # noqa: E712
        ).all()
        total_purchases = sum((p.amount_aed or Decimal('0') for p in purchases), Decimal('0'))
        total_paid = sum((p.amount_aed or Decimal('0') for p in payments), Decimal('0'))
        balance = total_purchases - total_paid
        if balance <= 0:
            continue
        oldest = min((p.purchase_date for p in purchases if p.purchase_date), default=None)
        days_old = (now - oldest.replace(tzinfo=tz.utc)).days if oldest else 0
        total_payable += balance
        rows.append([supplier.name, float(total_purchases), float(total_paid),
                     float(balance), days_old, supplier.payment_terms_days or 30])
    summary = {'إجمالي الذمم الدائنة': float(total_payable)}

    def build():
        if fmt == 'pdf':
            return ExportService.export_to_pdf('أعمار الذمم الدائنة — AP Aging',
                                               headers, rows, summary=summary)
        elif fmt == 'csv':
            return ExportService.export_to_csv(rows, headers, 'ap_aging.csv')
        return ExportService.export_to_excel('أعمار الذمم الدائنة — AP Aging',
                                             headers, rows, summary=summary)

    ext = {'xlsx': 'xlsx', 'pdf': 'pdf', 'csv': 'csv'}.get(fmt, 'xlsx')
    return _send_export(build, f'ap_aging.{ext}', ext)


@reports_bp.route('/cash-flow/export')
@login_required
@permission_required('view_reports')
def export_cash_flow():
    from services.export_service import ExportService
    from models import GLAccount, GLJournalLine, GLJournalEntry
    from sqlalchemy import func as sa_func
    fmt = request.args.get('format', 'xlsx')
    date_from = request.args.get('date_from', '', type=str)
    date_to = request.args.get('date_to', '', type=str)

    line_query = db.session.query(
        GLAccount.type.label('account_type'),
        GLAccount.code.label('account_code'),
        GLAccount.name.label('account_name'),
        GLAccount.name_ar.label('account_name_ar'),
        sa_func.sum(GLJournalLine.debit).label('debit'),
        sa_func.sum(GLJournalLine.credit).label('credit'),
    ).join(
        GLJournalEntry, GLJournalLine.entry_id == GLJournalEntry.id
    ).join(
        GLAccount, GLJournalLine.account_id == GLAccount.id
    ).filter(GLJournalEntry.is_posted == True, GLJournalEntry.is_reversed == False)  # noqa: E712

    if date_from:
        line_query = line_query.filter(sa_func.date(GLJournalEntry.entry_date) >= date_from)
    if date_to:
        line_query = line_query.filter(sa_func.date(GLJournalEntry.entry_date) <= date_to)

    lines = line_query.group_by(GLAccount.type, GLAccount.code, GLAccount.name, GLAccount.name_ar).all()
    headers = ['القسم', 'الحساب', 'المدين', 'الدائن', 'صافي']
    rows = []
    operating = investing = financing = Decimal('0')
    for line in lines:
        net = (line.debit or Decimal('0')) - (line.credit or Decimal('0'))
        cat = 'تشغيلي' if line.account_type == 'revenue' else (
              'استثماري' if line.account_type == 'asset' else 'تمويلي')
        if line.account_type == 'revenue':
            net = -net
            operating += net
        elif line.account_type == 'expense':
            operating += net
        elif line.account_type == 'asset':
            investing += net
        else:
            financing -= net
        rows.append([cat, f"{line.account_code} - {line.account_name_ar or line.account_name}",
                     float(line.debit or 0), float(line.credit or 0), float(net)])
    summary = {'تشغيلي': float(operating), 'استثماري': float(investing),
               'تمويلي': float(financing), 'الصافي': float(operating + investing + financing)}

    def build():
        if fmt == 'pdf':
            return ExportService.export_to_pdf('قائمة التدفقات النقدية — Cash Flow',
                                               headers, rows, summary=summary)
        elif fmt == 'csv':
            return ExportService.export_to_csv(rows, headers, 'cash_flow.csv')
        return ExportService.export_to_excel('قائمة التدفقات النقدية — Cash Flow',
                                             headers, rows, summary=summary)

    ext = {'xlsx': 'xlsx', 'pdf': 'pdf', 'csv': 'csv'}.get(fmt, 'xlsx')
    return _send_export(build, f'cash_flow.{ext}', ext)


@reports_bp.route('/vat-report/export')
@login_required
@permission_required('view_reports')
def export_vat_report():  # noqa: C901
    from services.export_service import ExportService
    from models import Sale, Purchase
    fmt = request.args.get('format', 'xlsx')
    date_from = request.args.get('date_from', '', type=str)
    date_to = request.args.get('date_to', '', type=str)

    # Sales
    sq = Sale.query.filter_by(status='confirmed')
    if date_from:
        sq = sq.filter(func.date(Sale.sale_date) >= date_from)
    if date_to:
        sq = sq.filter(func.date(Sale.sale_date) <= date_to)
    sales = sq.all()
    sales_headers = ['رقم الفاتورة', 'التاريخ', 'العميل', 'المبلغ', 'نسبة الضريبة', 'الضريبة']
    sales_rows = []
    total_output = Decimal('0')
    for s in sales:
        tr = s.tax_rate or Decimal('0')
        ta = s.tax_amount or Decimal('0')
        if tr > 0:
            sales_rows.append([s.sale_number, s.sale_date.strftime('%Y-%m-%d') if s.sale_date else '',
                               s.customer.name if s.customer else 'عميل نقدي',
                               float(s.amount_aed or 0), f'{tr}%', float(ta)])
            total_output += ta

    # Purchases
    pq = Purchase.query.filter_by(status='confirmed')
    if date_from:
        pq = pq.filter(func.date(Purchase.purchase_date) >= date_from)
    if date_to:
        pq = pq.filter(func.date(Purchase.purchase_date) <= date_to)
    purchases = pq.all()
    purchase_headers = ['رقم الفاتورة', 'التاريخ', 'المورد', 'المبلغ', 'نسبة الضريبة', 'الضريبة']
    purchase_rows = []
    total_input = Decimal('0')
    for p in purchases:
        tr = p.tax_rate or Decimal('0')
        ta = p.tax_amount or Decimal('0')
        if tr > 0:
            purchase_rows.append([p.purchase_number, p.purchase_date.strftime('%Y-%m-%d') if p.purchase_date else '',
                                  p.supplier.name if p.supplier else p.supplier_name,
                                  float(p.amount_aed or 0), f'{tr}%', float(ta)])
            total_input += ta

    net_vat = total_output - total_input

    def build_xlsx():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        # Sales sheet
        ws = wb.active
        ws.title = 'المبيعات الخاضعة'
        ws.sheet_view.rightToLeft = True
        hf = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
        for ci, h in enumerate(sales_headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hf
            c.fill = hfill
        for ri, row in enumerate(sales_rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        ws.cell(row=len(sales_rows) + 2, column=3, value='المجموع').font = Font(bold=True)
        ws.cell(row=len(sales_rows) + 2, column=6, value=float(total_output)).font = Font(bold=True)

        # Purchases sheet
        ws2 = wb.create_sheet('المشتريات الخاضعة')
        ws2.sheet_view.rightToLeft = True
        hfill2 = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
        for ci, h in enumerate(purchase_headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = hf
            c.fill = hfill2
        for ri, row in enumerate(purchase_rows, 2):
            for ci, val in enumerate(row, 1):
                ws2.cell(row=ri, column=ci, value=val)
        ws2.cell(row=len(purchase_rows) + 2, column=3, value='المجموع').font = Font(bold=True)
        ws2.cell(row=len(purchase_rows) + 2, column=6, value=float(total_input)).font = Font(bold=True)

        # Summary sheet
        ws3 = wb.create_sheet('الملخص')
        ws3.sheet_view.rightToLeft = True
        ws3.cell(row=1, column=1, value='ضريبة المبيعات (Output VAT)').font = Font(bold=True)
        ws3.cell(row=1, column=2, value=float(total_output))
        ws3.cell(row=2, column=1, value='ضريبة المشتريات (Input VAT)').font = Font(bold=True)
        ws3.cell(row=2, column=2, value=float(total_input))
        ws3.cell(row=3, column=1, value='الرصيد النهائي').font = Font(bold=True, size=12)
        ws3.cell(row=3, column=2, value=float(net_vat)).font = Font(bold=True, size=12)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def build():
        if fmt == 'pdf':
            # Combine both into one PDF
            all_rows = [['── المبيعات ──', '', '', '', '', '']] + sales_rows + \
                       [['', '', '', '', '', '']] + \
                       [['── المشتريات ──', '', '', '', '', '']] + purchase_rows
            summary = {'Output VAT': float(total_output), 'Input VAT': float(total_input),
                       'Net VAT': float(net_vat)}
            return ExportService.export_to_pdf('تقرير VAT — UAE 5%',
                                               sales_headers, all_rows, summary=summary)
        elif fmt == 'csv':
            combined = sales_rows + purchase_rows
            return ExportService.export_to_csv(combined, sales_headers, 'vat_report.csv')
        return build_xlsx()

    ext = {'xlsx': 'xlsx', 'pdf': 'pdf', 'csv': 'csv'}.get(fmt, 'xlsx')
    return _send_export(build, f'vat_report.{ext}', ext)
